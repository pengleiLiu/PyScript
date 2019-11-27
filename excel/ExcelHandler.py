import datetime
import time
import openpyxl
from openpyxl import load_workbook, Workbook


# openpyxl link：https://openpyxl.readthedocs.io/en/stable/usage.html#read-an-existing-workbook

# 读取excel表格 指定单元格数据
def read_work_book_row():
    wb = load_workbook(filename="/Users/xingshulin/Desktop/银行卡实名认证反馈.xlsx")
    auth_failed = wb['实名认证失败']

    print(auth_failed)
    print(auth_failed['B2'].value)


# 读取execl信息并遍历
def iter_work_sheet_row():
    wb = load_workbook(filename="/Users/xingshulin/Desktop/银行卡实名认证反馈.xlsx")
    ws = wb.active
    failed_sheet = wb['实名认证失败']
    print("数据表的最大行数" + str(failed_sheet.max_row))
    print("数据表的最大列数" + str(failed_sheet.max_column))
    print('=======开始遍历=========')
    DATA = []
    for row in failed_sheet.iter_rows(min_col=1, max_col=failed_sheet.max_column, min_row=1,
                                      max_row=failed_sheet.max_row):
        ROW = []
        for cell in row:
            if cell.row > 1 and cell.column == 4:
                ROW.append(str(cell.value))
            else:
                ROW.append(cell.value)
        DATA.append(ROW)
    print(DATA)


# 读取excel指定行 并根据规则进行修改
def process_data():
    # 读取表格内容
    wb = load_workbook(filename="/Users/xingshulin/Desktop/银行卡实名认证反馈.xlsx")
    sheet = wb['实名认证失败']

    data = [['姓名', '身份证号', '性别', '错误信息', '时间']]
    for row in sheet.iter_rows(min_col=1, max_col=sheet.max_column, min_row=2,
                               max_row=sheet.max_row):
        ROW = []
        for cell in row:
            if cell.row > 1 and cell.column == 4:
                ROW.append(str(cell.value))
            elif cell.column == 2:
                id_card = cell.value
                gender = get_gender(id_card[len(id_card) - 2])
                ROW.append(cell.value)
                ROW.append(gender)
            else:
                ROW.append(cell.value)
        data.append(ROW)
    print(data)

    return data


# 写入新的excel
def write_sheet():
    work_book_data = process_data()
    print(work_book_data)
    wb = Workbook()
    wb.active()
    dest_filename = '/Users/xingshulin/Desktop/test.xlsx'
    ws = wb.create_sheet("gender")
    for x in work_book_data:
        ws.append(x)

    wb.save(filename=dest_filename)


def get_gender(str):
    num = int(str)
    if (num % 2) == 0:
        return "女"
    else:
        return "男"


if __name__ == '__main__':
    # read_work_book_row()
    # iter_work_sheet_row()
    write_sheet()
